#!/usr/bin/env python3
"""
Bond Dissociation Energy (BDE) Analysis for First-Row Transition Metal Diatomics

This script analyzes CCSD(T)@DFA, KS-DFT, and ph-AFQMC bond dissociation energies
for 50 first-row transition metal diatomic molecules and generates publication-ready
Excel tables comparing computed values against experimental reference data.

Output: bde_results_<unit>.xlsx with 5 sheets:
    1. CC         - CCSD(T)@DFA results by bond type
    2. KS-DFT     - Kohn-Sham DFT results by bond type  
    3. ph-AFQMC   - Phaseless AFQMC results by bond type
    4. Overall M-H, M-O, M-Cl - Combined statistics (30 species)
    5. Overall M-H+, M-M      - Combined statistics (20 species)

Required files:
    - species/          Directory with ORCA output files
    - ref_data/species_ref_data.csv    Experimental BDE values
    - ref_data/ph-afqmc_data.csv       ph-AFQMC computed values

Usage:
    python generate_tables.py

Dependencies:
    numpy, pandas, scipy, openpyxl (auto-installed if missing)

Author: Barbaro Zulueta
Date: December 2025
"""

import sys
import subprocess
from pathlib import Path
from typing import Optional, Dict, List, Tuple

# Auto-install openpyxl if needed
try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])

import numpy as np
import pandas as pd
from scipy import constants
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =============================================================================
# CONFIGURATION
# =============================================================================

SPECIES_DIR = Path('species')
REF_DATA_FILE = Path('ref_data/species_ref_data.csv')
QMC_DATA_FILE = Path('ref_data/ph-afqmc_data.csv')

# Methods
CC_METHODS = ['ccsdt_hf', 'ccsdt_svwn5', 'ccsdt_pbe', 'ccsdt_pw91', 'ccsdt_r2scan', 'ccsdt_pbe0']
DFT_METHODS = ['svwn5', 'pbe', 'pw91', 'r2scan', 'b3lyp', 'pbe0', 'wb97x-v', 'wb97m-v']
QMC_METHODS = ['qmc_mp2', 'qmc_cc', 'qmc_tz_qz']

METHOD_LABELS = {
    'ccsdt_hf': 'CC-HF', 'ccsdt_svwn5': 'CC-SVWN5', 'ccsdt_pbe': 'CC-PBE',
    'ccsdt_pw91': 'CC-PW91', 'ccsdt_r2scan': 'CC-R2SCAN', 'ccsdt_pbe0': 'CC-PBE0',
    'svwn5': 'SVWN5', 'pbe': 'PBE', 'pw91': 'PW91', 'r2scan': 'R2SCAN',
    'b3lyp': 'B3LYP', 'pbe0': 'PBE0', 'wb97x-v': 'ωB97X-V', 'wb97m-v': 'ωB97M-V',
}
QMC_LABELS = {'qmc_mp2': 'QMC-MP2', 'qmc_cc': 'QMC-CC', 'qmc_tz_qz': 'QMC-TZ/QZ'}

BOND_TYPES = ['M-H', 'M-H+', 'M-O', 'M-Cl', 'M-M']
QMC_BOND_TYPES = ['M-H', 'M-O', 'M-Cl']  # QMC data available for most of these only


# =============================================================================
# UNIT CONVERSION (using CODATA constants)
# =============================================================================

_hartree_J = constants.value('Hartree energy')
_avogadro = constants.Avogadro
_calorie = constants.calorie

HARTREE_TO_KCAL = _hartree_J * _avogadro / (_calorie * 1000)
HARTREE_TO_EV = constants.value('Hartree energy in eV')
KCAL_TO_EV = HARTREE_TO_EV / HARTREE_TO_KCAL
KCAL_TO_KJ = _calorie

UNIT_CONVERSION = {'kcal/mol': 1.0, 'kJ/mol': KCAL_TO_KJ, 'eV': KCAL_TO_EV}

# Global state
ENERGY_UNIT = 'eV'
REF_DATA: Dict = {}
SPECIES_BY_TYPE: Dict[str, List[str]] = {}
QMC_DATA: Dict = {}
cc_results: Dict = {}
dft_results: Dict = {}


def convert_energy(value_kcal: float) -> float:
    """Convert energy from kcal/mol to selected unit."""
    return value_kcal * UNIT_CONVERSION[ENERGY_UNIT]


# =============================================================================
# DATA LOADING
# =============================================================================

def classify_species(name: str) -> Tuple[str, List[str]]:
    """Classify species by bond type and return constituent atoms."""
    name = name.strip()
    if name.endswith('+'):
        metal = name.replace('-H+', '')
        return 'M-H+', [f"{metal}+", 'H']
    parts = name.split('-')
    atom1, atom2 = parts[0], parts[1]
    bond_types = {'H': 'M-H', 'O': 'M-O', 'Cl': 'M-Cl'}
    if atom2 in bond_types:
        return bond_types[atom2], [atom1, atom2]
    if atom1 == atom2:
        return 'M-M', [atom1, atom2]
    return 'Other', [atom1, atom2]


def load_data() -> None:
    """Load experimental reference data and ph-AFQMC results."""
    global REF_DATA, SPECIES_BY_TYPE, QMC_DATA
    
    # Load experimental reference data
    df_ref = pd.read_csv(REF_DATA_FILE)
    df_ref['dimers'] = df_ref['dimers'].str.strip()
    df_ref[['bond_type', 'atoms']] = df_ref['dimers'].apply(
        lambda x: pd.Series(classify_species(x))
    )
    df_ref['E_SO'] = df_ref['E_SO'].fillna(0.0)
    df_ref['uncertainty'] = df_ref['uncertainty'].fillna(0.0)
    
    REF_DATA = df_ref.set_index('dimers').to_dict('index')
    SPECIES_BY_TYPE = {
        bt: df_ref[df_ref['bond_type'] == bt]['dimers'].tolist() 
        for bt in BOND_TYPES
    }
    
    # Load ph-AFQMC data
    df_qmc = pd.read_csv(QMC_DATA_FILE)
    df_qmc['dimers'] = df_qmc['dimers'].str.strip()
    df_qmc = df_qmc.rename(columns={
        'qmc_tz/qz': 'qmc_tz_qz',
        'qm_tz/qz_uncertainty': 'qmc_tz_qz_uncertainty',
        'qm_cc_uncertainty': 'qmc_cc_uncertainty',
    })
    
    QMC_DATA = {}
    for _, row in df_qmc.iterrows():
        sp = row['dimers']
        QMC_DATA[sp] = {}
        for m in QMC_METHODS:
            val = row.get(m)
            unc = row.get(f'{m}_uncertainty')
            if pd.notna(val):
                QMC_DATA[sp][m] = (val, unc if pd.notna(unc) else 0.0)


# =============================================================================
# ORCA OUTPUT PARSING
# =============================================================================

def parse_cc_energy(path: Path) -> Optional[float]:
    """Parse CCSD(T) energy from CBS extrapolation block in ORCA output."""
    if not path.exists():
        return None
    energy = None
    in_extrapolation = False
    with open(path) as f:
        for line in f:
            if 'Extrapolation of energy' in line:
                in_extrapolation = True
            elif in_extrapolation and 'FINAL SINGLE POINT ENERGY' in line:
                energy = float(line.split()[-1])
                break
    return energy


def parse_dft_energy(path: Path) -> Optional[float]:
    """Parse DFT energy (last FINAL SINGLE POINT ENERGY) from ORCA output."""
    if not path.exists():
        return None
    energy = None
    with open(path) as f:
        for line in f:
            if 'FINAL SINGLE POINT ENERGY' in line:
                energy = float(line.split()[-1])
    return energy


def get_file_path(name: str, method: str, parent_species: str = None) -> Path:
    """
    Construct path to ORCA output file.
    
    Handles special cases:
    - H atom uses HF reference for CC methods
    - Mn-Mn uses ZORA instead of X2C for CC-PBE0
    """
    is_cc = method.startswith('ccsdt')
    
    # Special case: H atom always uses HF for CC
    if name == 'H' and is_cc:
        return SPECIES_DIR / 'H' / 'H_hf_cbs_x2c.out'
    
    # Special case: Mn-Mn uses ZORA for CC-PBE0
    use_zora = (method == 'ccsdt_pbe0') and (parent_species == 'Mn-Mn' or name == 'Mn-Mn')
    rela = 'zora' if use_zora else 'x2c'
    basis = 'cbs' if is_cc else 'tz'
    
    return SPECIES_DIR / name / f"{name}_{method}_{basis}_{rela}.out"


# =============================================================================
# BDE CALCULATION
# =============================================================================

def calculate_bde(species: str, method: str) -> Optional[float]:
    """
    Calculate bond dissociation energy for a species using specified method.
    
    BDE = E(atoms) - E(molecule) + E_SO
    
    Returns BDE in kcal/mol, or None if calculation files are missing.
    """
    is_cc = method.startswith('ccsdt')
    parse_func = parse_cc_energy if is_cc else parse_dft_energy
    
    # Parse molecular energy
    mol_path = get_file_path(species, method, parent_species=species)
    e_mol = parse_func(mol_path)
    if e_mol is None:
        return None
    
    # Parse atomic energies
    atoms = REF_DATA[species]['atoms']
    e_atoms = 0.0
    for atom in atoms:
        atom_path = get_file_path(atom, method, parent_species=species)
        e_atom = parse_func(atom_path)
        if e_atom is None:
            return None
        e_atoms += e_atom
    
    # Calculate BDE with spin-orbit correction
    e_so = REF_DATA[species]['E_SO']
    bde_hartree = e_atoms - e_mol
    bde_kcal = bde_hartree * HARTREE_TO_KCAL + e_so
    
    return bde_kcal


def calculate_all_bdes() -> None:
    """Calculate BDEs for all species using all CC and DFT methods."""
    global cc_results, dft_results
    
    cc_results = {sp: {} for sp in REF_DATA}
    dft_results = {sp: {} for sp in REF_DATA}
    
    for species in REF_DATA:
        for method in CC_METHODS:
            bde = calculate_bde(species, method)
            if bde is not None:
                cc_results[species][method] = bde
        
        for method in DFT_METHODS:
            bde = calculate_bde(species, method)
            if bde is not None:
                dft_results[species][method] = bde


# =============================================================================
# STATISTICS
# =============================================================================

def calc_rmse(errors: np.ndarray) -> float:
    """Root mean square error."""
    return float(np.sqrt(np.mean(errors**2)))

def calc_mae(errors: np.ndarray) -> float:
    """Mean absolute error."""
    return float(np.mean(np.abs(errors)))

def calc_max(errors: np.ndarray) -> float:
    """Maximum absolute error."""
    return float(np.max(np.abs(errors)))


# =============================================================================
# EXCEL GENERATION
# =============================================================================

def generate_excel(unit_str: str) -> str:
    """
    Generate Excel workbook with BDE analysis results.
    
    Creates 5 sheets:
    1. CC - CCSD(T)@DFA results with CC-Best column
    2. KS-DFT - DFT results
    3. ph-AFQMC - QMC results with QMC-Best column
    4. Overall M-H, M-O, M-Cl - Summary statistics
    5. Overall M-H+, M-M - Summary statistics
    """
    wb = Workbook()
    
    # Excel styles
    STYLES = {
        'header_font': Font(bold=True),
        'title_font': Font(bold=True, size=12),
        'section_font': Font(bold=True, italic=True),
        'header_fill': PatternFill('solid', fgColor='D9E1F2'),
        'cc_best_fill': PatternFill('solid', fgColor='E2EFDA'),    # Green
        'qmc_best_fill': PatternFill('solid', fgColor='FFF2CC'),   # Yellow/Gold
        'center': Alignment(horizontal='center'),
        'border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
    }
    
    # -------------------------------------------------------------------------
    # Helper functions
    # -------------------------------------------------------------------------
    
    def set_cell(ws, row, col, value='', fill=None):
        """Set cell value with border and optional fill."""
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = STYLES['border']
        if fill:
            cell.fill = fill
        return cell
    
    def write_headers(ws, row, headers):
        """Write styled header row."""
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = STYLES['header_font']
            cell.fill = STYLES['header_fill']
            cell.alignment = STYLES['center']
            cell.border = STYLES['border']
    
    def set_col_widths(ws, num_cols, width=12):
        """Set uniform column widths."""
        for col in range(1, num_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def write_legend(ws, row, items):
        """Write color legend."""
        ws.cell(row=row, column=1, value='Legend:').font = STYLES['header_font']
        for fill, text in items:
            row += 1
            ws.cell(row=row, column=1, value='').fill = fill
            ws.cell(row=row, column=2, value=text)
    
    def get_method_error(sp, method, results):
        """Get BDE value and error vs experiment for a species/method."""
        bde_kcal = results.get(sp, {}).get(method)
        if bde_kcal is None:
            return None, None
        exp = convert_energy(REF_DATA[sp]['D_e'])
        bde = convert_energy(bde_kcal)
        return bde, bde - exp
    
    def find_best_index(values_errors):
        """Find index of method with lowest absolute error."""
        best_idx, best_err = None, None
        for i, (_, err) in enumerate(values_errors):
            if err is not None and (best_err is None or abs(err) < abs(best_err)):
                best_idx, best_err = i, err
        return best_idx
    
    # -------------------------------------------------------------------------
    # Sheet 1: CC (CCSD(T)@DFA)
    # -------------------------------------------------------------------------
    
    ws1 = wb.active
    ws1.title = 'CC'
    
    headers_cc = ['Species', f'Exp ({ENERGY_UNIT})']
    for m in CC_METHODS:
        headers_cc.extend([METHOD_LABELS[m], f'{METHOD_LABELS[m]} Error'])
    headers_cc.extend(['CC-Best', 'CC-Best Error'])
    
    row = 1
    for bond_type in BOND_TYPES:
        # Table title
        ws1.cell(row=row, column=1, value=f'{bond_type} Bond Type').font = STYLES['title_font']
        row += 1
        write_headers(ws1, row, headers_cc)
        row += 1
        
        errors = {m: [] for m in CC_METHODS}
        errors['cc_best'] = []
        
        # Data rows
        for sp in SPECIES_BY_TYPE[bond_type]:
            exp = convert_energy(REF_DATA[sp]['D_e'])
            set_cell(ws1, row, 1, sp)
            set_cell(ws1, row, 2, round(exp, 4))
            
            # Compute values and find best
            vals_errs = [get_method_error(sp, m, cc_results) for m in CC_METHODS]
            best_idx = find_best_index(vals_errs)
            
            col = 3
            for i, m in enumerate(CC_METHODS):
                bde, err = vals_errs[i]
                if bde is not None:
                    errors[m].append(err)
                    fill = STYLES['cc_best_fill'] if i == best_idx else None
                    set_cell(ws1, row, col, round(bde, 4), fill)
                    set_cell(ws1, row, col+1, round(err, 4), fill)
                else:
                    set_cell(ws1, row, col)
                    set_cell(ws1, row, col+1)
                col += 2
            
            # CC-Best column
            if best_idx is not None:
                bde, err = vals_errs[best_idx]
                errors['cc_best'].append(err)
                set_cell(ws1, row, col, round(bde, 4), STYLES['cc_best_fill'])
                set_cell(ws1, row, col+1, round(err, 4), STYLES['cc_best_fill'])
            else:
                set_cell(ws1, row, col)
                set_cell(ws1, row, col+1)
            row += 1
        
        # Statistics rows
        for stat_name, stat_func in [('RMSE', calc_rmse), ('MAE', calc_mae), ('MAX', calc_max)]:
            set_cell(ws1, row, 1, stat_name).font = STYLES['header_font']
            set_cell(ws1, row, 2)
            col = 3
            for m in CC_METHODS:
                set_cell(ws1, row, col)
                val = round(stat_func(np.array(errors[m])), 4) if errors[m] else ''
                set_cell(ws1, row, col+1, val)
                col += 2
            set_cell(ws1, row, col)
            val = round(stat_func(np.array(errors['cc_best'])), 4) if errors['cc_best'] else ''
            set_cell(ws1, row, col+1, val, STYLES['cc_best_fill'])
            row += 1
        
        # N row
        set_cell(ws1, row, 1, 'N').font = STYLES['header_font']
        set_cell(ws1, row, 2)
        col = 3
        for m in CC_METHODS:
            set_cell(ws1, row, col)
            set_cell(ws1, row, col+1, len(errors[m]) if errors[m] else '')
            col += 2
        set_cell(ws1, row, col)
        set_cell(ws1, row, col+1, len(errors['cc_best']) if errors['cc_best'] else '', STYLES['cc_best_fill'])
        row += 3  # Gap between tables
    
    write_legend(ws1, row, [(STYLES['cc_best_fill'], 'CC-Best: Best CCSD(T) method (lowest |error|)')])
    set_col_widths(ws1, len(headers_cc))
    
    # -------------------------------------------------------------------------
    # Sheet 2: KS-DFT
    # -------------------------------------------------------------------------
    
    ws2 = wb.create_sheet('KS-DFT')
    
    headers_dft = ['Species', f'Exp ({ENERGY_UNIT})']
    for m in DFT_METHODS:
        headers_dft.extend([METHOD_LABELS[m], f'{METHOD_LABELS[m]} Error'])
    
    row = 1
    for bond_type in BOND_TYPES:
        ws2.cell(row=row, column=1, value=f'{bond_type} Bond Type').font = STYLES['title_font']
        row += 1
        write_headers(ws2, row, headers_dft)
        row += 1
        
        errors = {m: [] for m in DFT_METHODS}
        
        for sp in SPECIES_BY_TYPE[bond_type]:
            exp = convert_energy(REF_DATA[sp]['D_e'])
            set_cell(ws2, row, 1, sp)
            set_cell(ws2, row, 2, round(exp, 4))
            
            col = 3
            for m in DFT_METHODS:
                bde, err = get_method_error(sp, m, dft_results)
                if bde is not None:
                    errors[m].append(err)
                    set_cell(ws2, row, col, round(bde, 4))
                    set_cell(ws2, row, col+1, round(err, 4))
                else:
                    set_cell(ws2, row, col)
                    set_cell(ws2, row, col+1)
                col += 2
            row += 1
        
        # Statistics
        for stat_name, stat_func in [('RMSE', calc_rmse), ('MAE', calc_mae), ('MAX', calc_max)]:
            set_cell(ws2, row, 1, stat_name).font = STYLES['header_font']
            set_cell(ws2, row, 2)
            col = 3
            for m in DFT_METHODS:
                set_cell(ws2, row, col)
                val = round(stat_func(np.array(errors[m])), 4) if errors[m] else ''
                set_cell(ws2, row, col+1, val)
                col += 2
            row += 1
        
        # N row
        set_cell(ws2, row, 1, 'N').font = STYLES['header_font']
        set_cell(ws2, row, 2)
        col = 3
        for m in DFT_METHODS:
            set_cell(ws2, row, col)
            set_cell(ws2, row, col+1, len(errors[m]) if errors[m] else '')
            col += 2
        row += 3
    
    set_col_widths(ws2, len(headers_dft))
    
    # -------------------------------------------------------------------------
    # Sheet 3: ph-AFQMC
    # -------------------------------------------------------------------------
    
    ws3 = wb.create_sheet('ph-AFQMC')
    
    headers_qmc = ['Species', f'Exp ({ENERGY_UNIT})']
    for m in QMC_METHODS:
        headers_qmc.extend([QMC_LABELS[m], f'{QMC_LABELS[m]} Unc', f'{QMC_LABELS[m]} Error'])
    headers_qmc.extend(['QMC-Best', 'QMC-Best Unc', 'QMC-Best Error'])
    
    row = 1
    for bond_type in QMC_BOND_TYPES:
        ws3.cell(row=row, column=1, value=f'{bond_type} Bond Type').font = STYLES['title_font']
        row += 1
        write_headers(ws3, row, headers_qmc)
        row += 1
        
        errors = {m: [] for m in QMC_METHODS}
        errors['qmc_best'] = []
        
        for sp in SPECIES_BY_TYPE[bond_type]:
            if sp not in QMC_DATA:
                continue
            
            exp = convert_energy(REF_DATA[sp]['D_e'])
            set_cell(ws3, row, 1, sp)
            set_cell(ws3, row, 2, round(exp, 4))
            
            # Collect QMC values
            vals = []
            for m in QMC_METHODS:
                if m in QMC_DATA[sp]:
                    val_kcal, unc_kcal = QMC_DATA[sp][m]
                    val, unc = convert_energy(val_kcal), convert_energy(unc_kcal)
                    vals.append((val, unc, val - exp))
                else:
                    vals.append((None, None, None))
            
            # Find best
            best_idx = None
            for i, (_, _, err) in enumerate(vals):
                if err is not None and (best_idx is None or abs(err) < abs(vals[best_idx][2])):
                    best_idx = i
            
            col = 3
            for i, m in enumerate(QMC_METHODS):
                val, unc, err = vals[i]
                if val is not None:
                    errors[m].append(err)
                    fill = STYLES['qmc_best_fill'] if i == best_idx else None
                    set_cell(ws3, row, col, round(val, 4), fill)
                    set_cell(ws3, row, col+1, round(unc, 4), fill)
                    set_cell(ws3, row, col+2, round(err, 4), fill)
                else:
                    set_cell(ws3, row, col)
                    set_cell(ws3, row, col+1)
                    set_cell(ws3, row, col+2)
                col += 3
            
            # QMC-Best column
            if best_idx is not None:
                val, unc, err = vals[best_idx]
                errors['qmc_best'].append(err)
                set_cell(ws3, row, col, round(val, 4), STYLES['qmc_best_fill'])
                set_cell(ws3, row, col+1, round(unc, 4), STYLES['qmc_best_fill'])
                set_cell(ws3, row, col+2, round(err, 4), STYLES['qmc_best_fill'])
            else:
                set_cell(ws3, row, col)
                set_cell(ws3, row, col+1)
                set_cell(ws3, row, col+2)
            row += 1
        
        # Statistics
        for stat_name, stat_func in [('RMSE', calc_rmse), ('MAE', calc_mae), ('MAX', calc_max)]:
            set_cell(ws3, row, 1, stat_name).font = STYLES['header_font']
            set_cell(ws3, row, 2)
            col = 3
            for m in QMC_METHODS:
                set_cell(ws3, row, col)
                set_cell(ws3, row, col+1)
                val = round(stat_func(np.array(errors[m])), 4) if errors[m] else ''
                set_cell(ws3, row, col+2, val)
                col += 3
            set_cell(ws3, row, col)
            set_cell(ws3, row, col+1)
            val = round(stat_func(np.array(errors['qmc_best'])), 4) if errors['qmc_best'] else ''
            set_cell(ws3, row, col+2, val, STYLES['qmc_best_fill'])
            row += 1
        
        # N row
        set_cell(ws3, row, 1, 'N').font = STYLES['header_font']
        set_cell(ws3, row, 2)
        col = 3
        for m in QMC_METHODS:
            set_cell(ws3, row, col)
            set_cell(ws3, row, col+1)
            set_cell(ws3, row, col+2, len(errors[m]) if errors[m] else '')
            col += 3
        set_cell(ws3, row, col)
        set_cell(ws3, row, col+1)
        set_cell(ws3, row, col+2, len(errors['qmc_best']) if errors['qmc_best'] else '', STYLES['qmc_best_fill'])
        row += 3
    
    write_legend(ws3, row, [(STYLES['qmc_best_fill'], 'QMC-Best: Best ph-AFQMC method (lowest |error|)')])
    set_col_widths(ws3, len(headers_qmc))
    
    # -------------------------------------------------------------------------
    # Sheets 4 & 5: Overall Statistics
    # -------------------------------------------------------------------------
    
    headers_overall = ['Method', f'RMSE ({ENERGY_UNIT})', f'MAE ({ENERGY_UNIT})', f'MAX ({ENERGY_UNIT})', 'N']
    
    def collect_errors(bond_types, methods, results):
        """Collect errors across bond types for given methods."""
        all_errors = {m: [] for m in methods}
        for bt in bond_types:
            for sp in SPECIES_BY_TYPE[bt]:
                for m in methods:
                    _, err = get_method_error(sp, m, results)
                    if err is not None:
                        all_errors[m].append(err)
        return all_errors
    
    def collect_best_errors(bond_types, methods, results):
        """Collect best errors across bond types."""
        best_errors = []
        for bt in bond_types:
            for sp in SPECIES_BY_TYPE[bt]:
                exp = convert_energy(REF_DATA[sp]['D_e'])
                best_err = None
                for m in methods:
                    bde_kcal = results.get(sp, {}).get(m)
                    if bde_kcal is not None:
                        err = convert_energy(bde_kcal) - exp
                        if best_err is None or abs(err) < abs(best_err):
                            best_err = err
                if best_err is not None:
                    best_errors.append(best_err)
        return best_errors
    
    def write_overall_sheet(ws, title, bond_types, include_qmc=False):
        """Write overall statistics sheet."""
        ws.cell(row=1, column=1, value=title).font = STYLES['title_font']
        write_headers(ws, 2, headers_overall)
        
        row = 3
        
        def write_section(title_text):
            nonlocal row
            cell = ws.cell(row=row, column=1, value=title_text)
            cell.font = STYLES['section_font']
            cell.alignment = STYLES['center']
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            row += 1
        
        def write_method_row(label, errs, fill=None):
            nonlocal row
            if not errs:
                return
            err_arr = np.array(errs)
            for c in range(1, 6):
                cell = ws.cell(row=row, column=c)
                cell.border = STYLES['border']
                if fill:
                    cell.fill = fill
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=round(calc_rmse(err_arr), 4))
            ws.cell(row=row, column=3, value=round(calc_mae(err_arr), 4))
            ws.cell(row=row, column=4, value=round(calc_max(err_arr), 4))
            ws.cell(row=row, column=5, value=len(err_arr))
            row += 1
        
        # CCSD(T) section
        write_section('CCSD(T)')
        for m in CC_METHODS:
            errs = collect_errors(bond_types, [m], cc_results)[m]
            write_method_row(METHOD_LABELS[m], errs)
        cc_best = collect_best_errors(bond_types, CC_METHODS, cc_results)
        write_method_row('CC-Best', cc_best, STYLES['cc_best_fill'])
        
        # KS-DFT section
        row += 1
        write_section('KS-DFT')
        for m in DFT_METHODS:
            errs = collect_errors(bond_types, [m], dft_results)[m]
            write_method_row(METHOD_LABELS[m], errs)
        
        # ph-AFQMC section
        if include_qmc:
            row += 1
            write_section('ph-AFQMC')
            for m in QMC_METHODS:
                errs = []
                for bt in bond_types:
                    for sp in SPECIES_BY_TYPE[bt]:
                        if sp in QMC_DATA and m in QMC_DATA[sp]:
                            val_kcal, _ = QMC_DATA[sp][m]
                            errs.append(convert_energy(val_kcal) - convert_energy(REF_DATA[sp]['D_e']))
                write_method_row(QMC_LABELS[m], errs)
            
            # QMC-Best
            qmc_best = []
            for bt in bond_types:
                for sp in SPECIES_BY_TYPE[bt]:
                    if sp not in QMC_DATA:
                        continue
                    exp = convert_energy(REF_DATA[sp]['D_e'])
                    best_err = None
                    for m in QMC_METHODS:
                        if m in QMC_DATA[sp]:
                            val_kcal, _ = QMC_DATA[sp][m]
                            err = convert_energy(val_kcal) - exp
                            if best_err is None or abs(err) < abs(best_err):
                                best_err = err
                    if best_err is not None:
                        qmc_best.append(best_err)
            write_method_row('QMC-Best', qmc_best, STYLES['qmc_best_fill'])
        
        # Legend
        legend_items = [(STYLES['cc_best_fill'], 'CC-Best: Best CCSD(T) method')]
        if include_qmc:
            legend_items.append((STYLES['qmc_best_fill'], 'QMC-Best: Best ph-AFQMC method'))
        write_legend(ws, row + 2, legend_items)
        set_col_widths(ws, len(headers_overall), width=14)
    
    # Sheet 4
    ws4 = wb.create_sheet('Overall M-H, M-O, M-Cl')
    write_overall_sheet(ws4, 'Overall Statistics: M-H, M-O, M-Cl Bond Types',
                        ['M-H', 'M-O', 'M-Cl'], include_qmc=True)
    
    # Sheet 5
    ws5 = wb.create_sheet('Overall M-H+, M-M')
    write_overall_sheet(ws5, 'Overall Statistics: M-H+, M-M Bond Types',
                        ['M-H+', 'M-M'], include_qmc=False)
    
    # Save workbook
    output_file = f'bde_results_{unit_str}.xlsx'
    wb.save(output_file)
    return output_file


# =============================================================================
# MAIN
# =============================================================================

def get_user_input() -> None:
    """Get energy unit selection from user."""
    global ENERGY_UNIT
    
    print("\nSelect energy unit:")
    print("  1. eV (default)")
    print("  2. kcal/mol")
    print("  3. kJ/mol")
    
    choice = input("Enter choice [1]: ").strip() or '1'
    
    if choice == '2':
        ENERGY_UNIT = 'kcal/mol'
    elif choice == '3':
        ENERGY_UNIT = 'kJ/mol'
    else:
        ENERGY_UNIT = 'eV'
    
    print(f"\nUsing energy unit: {ENERGY_UNIT}")


def main():
    """Main entry point."""
    get_user_input()
    
    print("\nLoading reference data...")
    load_data()
    
    print("Calculating BDEs...")
    calculate_all_bdes()
    
    cc_count = sum(len(v) for v in cc_results.values())
    dft_count = sum(len(v) for v in dft_results.values())
    print(f"Calculated {cc_count} CCSD(T)/CBS and {dft_count} DFT bond dissociation energies.")
    
    unit_str = ENERGY_UNIT.replace('/', '_')
    
    print("\nGenerating Excel file...")
    excel_file = generate_excel(unit_str)
    
    print(f"\n✓ Saved: {excel_file}")
    print("  Sheets:")
    print("    1. CC              - CCSD(T)@DFA/HF results by bond type")
    print("    2. KS-DFT          - DFT results by bond type")
    print("    3. ph-AFQMC        - QMC results by bond type")
    print("    4. Overall M-H, M-O, M-Cl  - Combined statistics (29 species)")
    print("    5. Overall M-H+, M-M       - Combined statistics (20 species)")
    print("  Color coding: Green = CC-Best, Yellow = QMC-Best")


if __name__ == '__main__':
    main()
